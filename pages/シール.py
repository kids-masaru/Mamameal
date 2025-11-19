import streamlit as st
import io
import os
from openpyxl import load_workbook
import pdfplumber

# --- ▼▼ 新しくインポート ▼▼ ---
import pytesseract
from pdf2image import convert_from_bytes
# --- ▲▲ 新しくインポート ▲▲ ---

# --- シール/その他PDF処理関数 ---
def process_other_pdf_to_seal_template(pdf_bytes_io, existing_seal_path):
    """
    seal.xlsxを読み込み、シートを2枚に分けてPDFデータを貼り付ける
    - 貼り付け1: OCR結果 (E列まで書き込み、F列以降は保護)
    - 貼り付け2: テキスト抽出 (M列まで書き込み、N列以降は保護)
    """
    wb = load_workbook(existing_seal_path)
    
    # --- シートの準備 ---
    ws1 = wb.worksheets[0]
    ws1.title = "貼り付け1"
    
    if "貼り付け2" in wb.sheetnames:
        ws2 = wb["貼り付け2"]
    else:
        ws2 = wb.create_sheet(title="貼り付け2", index=1)

    # --- ▼▼ 修正ポイント1：行削除をやめて、指定列のデータだけクリアする ▼▼ ---
    
    # 【貼り付け1】のクリア処理 (A列(1) 〜 E列(5) のみ消す)
    # max_rowまでの行をループし、1〜5列目の値をNoneにする
    if ws1.max_row > 0:
        for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.value = None

    # 【貼り付け2】のクリア処理 (A列(1) 〜 M列(13) のみ消す)
    # max_rowまでの行をループし、1〜13列目の値をNoneにする
    if ws2.max_row > 0:
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=13):
            for cell in row:
                cell.value = None
    
    # --- ▲▲ 修正ポイント1 完了 ▲▲ ---

    
    # --- ▼▼ 処理1: 【貼り付け1】OCRによる抽出 ▼▼ ---
    ws1_current_row = 1
    try:
        images = convert_from_bytes(pdf_bytes_io.getvalue())
        
        if not images:
            ws1.cell(row=1, column=1, value="PDFを画像に変換できませんでした。")
        else:
            ws1.cell(row=1, column=1, value="--- OCR抽出開始 ---")
            ws1_current_row += 1
            
            for i, page_image in enumerate(images, 1):
                ocr_text = pytesseract.image_to_string(page_image, lang='jpn', config='--psm 6')
                
                ws1.cell(row=ws1_current_row, column=1, value=f"--- ページ {i} (OCR) ---")
                ws1_current_row += 1
                
                if ocr_text:
                    lines = ocr_text.split('\n')
                    for line in lines:
                        if line.strip():
                            words = line.split() 
                            for col_idx, word in enumerate(words, 1):
                                if col_idx > 5: # E列(5)まで。F列以降は触らない
                                    break
                                ws1.cell(row=ws1_current_row, column=col_idx, value=word)
                            ws1_current_row += 1
                else:
                    ws1.cell(row=ws1_current_row, column=1, value="(認識不可)")
                    ws1_current_row += 1

    except Exception as e:
        ws1.cell(row=1, column=1, value=f"エラー: {str(e)}")
    # --- ▲▲ 処理1 完了 ▲▲ ---


    # --- ▼▼ 処理2: 【貼り付け2】従来のテキスト抽出 ▼▼ ---
    ws2_current_row = 1
    try:
        with pdfplumber.open(pdf_bytes_io) as pdf:
            if not pdf.pages:
                ws2.cell(row=1, column=1, value="PDFページなし")
            else:
                for page_number, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text()
                    
                    ws2.cell(row=ws2_current_row, column=1, value=f"--- ページ {page_number} (テキスト) ---")
                    ws2_current_row += 1
                    
                    if page_text:
                        lines = page_text.split('\n')
                        for line in lines:
                            if line.strip():
                                words = line.split()
                                for col_idx, word in enumerate(words, 1):
                                    if col_idx > 13: # M列(13)まで。N列以降は触らない
                                        break
                                    ws2.cell(row=ws2_current_row, column=col_idx, value=word)
                                ws2_current_row += 1
                    else:
                        ws2.cell(row=ws2_current_row, column=1, value="(抽出不可)")
                        ws2_current_row += 1
    
    except Exception as e:
        ws2.cell(row=1, column=1, value=f"エラー: {str(e)}")
        
    # --- ▲▲ 処理2 完了 ▲▲ ---

    # 変更を保存
    output_excel = io.BytesIO()
    wb.save(output_excel)
    return output_excel.getvalue()


# --- 以下、StreamlitのUI部分（変更なし） ---
st.markdown("""
    <style>
        [data-testid="stSidebarNav"] ul { display: none; }
        .custom-title {
            font-size: 2.1rem; font-weight: 600; color: #3A322E;
            padding-bottom: 10px; border-bottom: 3px solid #FF9933; margin-bottom: 25px;
        }
        .stApp { background: #fff5e6; }
    </style>
""", unsafe_allow_html=True)

st.sidebar.title("メニュー")
st.sidebar.page_link("streamlit_app.py", label="数出表 変換", icon="📄")
st.sidebar.page_link("pages/シール.py", label="シール貼付 変換", icon="🏷️")
st.sidebar.page_link("pages/マスタ設定.py", label="マスタ設定", icon="⚙️")
show_debug = st.sidebar.checkbox("デバッグ情報を表示", value=False)

st.markdown('<p class="custom-title">シール貼付 PDF変換ツール</p>', unsafe_allow_html=True)

uploaded_pdf = st.file_uploader("処理するシールPDFファイルをアップロードしてください", type="pdf", label_visibility="collapsed")

if uploaded_pdf is not None:
    pdf_bytes_io = io.BytesIO(uploaded_pdf.getvalue())
    
    st.info("📜 シール/その他PDFとして処理します...")
    seal_template_path = "seal.xlsx"
    
    if not os.path.exists(seal_template_path):
        st.error(f"必要なテンプレートファイル '{seal_template_path}' が見つかりません。")
        st.stop()
    
    try:
        with st.spinner(f"OCR処理とテキスト抽出を実行中..."):
            modified_seal_bytes = process_other_pdf_to_seal_template(pdf_bytes_io, seal_template_path)
        
        st.success(f"✅ 処理が完了しました！")
        
        original_pdf_name = os.path.splitext(uploaded_pdf.name)[0]
        
        st.download_button(
            label=f"▼ seal_{original_pdf_name}.xlsx ダウンロード",
            data=modified_seal_bytes,
            file_name=f"seal_{original_pdf_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"処理中にエラーが発生しました: {str(e)}")
        if show_debug:
            st.exception(e)
