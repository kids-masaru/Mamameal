import streamlit as st
import pandas as pd
import io
import os
import re
from openpyxl import load_workbook
import glob
import pdfplumber  # (新規) シール/その他PDFの読み取りに必要

from pdf_utils import (
    safe_write_df, pdf_to_excel_data_for_paste_sheet, extract_table_from_pdf_for_bento,
    find_correct_anchor_for_bento, extract_bento_range_for_bento, match_bento_data, 
    extract_detailed_client_info_from_pdf, export_detailed_client_data_to_dataframe
)

st.set_page_config(
    page_title="PDF変換ツール",
    page_icon="./static/icons/android-chrome-192.png",
    layout="centered",
)

def load_master_data(file_prefix, default_columns):
    list_of_files = glob.glob(os.path.join('.', f'{file_prefix}*.csv'))
    if not list_of_files:
        return pd.DataFrame(columns=default_columns)
    latest_file = max(list_of_files, key=os.path.getmtime)
    encodings = ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']
    for encoding in encodings:
        try:
            df = pd.read_csv(latest_file, encoding=encoding, dtype=str).fillna('')
            if not df.empty: return df
        except Exception:
            continue
    return pd.DataFrame(columns=default_columns)

def load_master_csv(file_pattern):
    """同じフォルダにあるCSVファイルからマスタデータを読み込む"""
    list_of_files = glob.glob(os.path.join('.', f'*{file_pattern}*.csv'))
    if not list_of_files:
        return pd.DataFrame()
    latest_file = max(list_of_files, key=os.path.getmtime)
    encodings = ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']
    for encoding in encodings:
        try:
            df = pd.read_csv(latest_file, encoding=encoding, dtype=str).fillna('')
            if not df.empty:
                df.columns = df.columns.str.strip()
                return df
        except Exception:
            continue
    return pd.DataFrame()

def paste_dataframe_to_sheet(ws, df, start_row=1, start_col=1):
    """DataFrameをExcelシートに貼り付ける"""
    # ヘッダーを貼り付け
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    
    # データを貼り付け
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=start_row + r_idx + 1, column=c_idx, value=value)

# --- 新規追加: シール/その他PDF処理関数 ---
def process_other_pdf_to_seal_template(pdf_bytes_io, existing_seal_path):
    """
    数出表以外のPDFを処理し、既存のseal.xlsxの最初のシートに全テーブルデータを貼り付ける
    """
    # 既存のseal.xlsxを読み込む
    wb = load_workbook(existing_seal_path)
    
    # 最初のシートを取得
    ws = wb.worksheets[0]
    
    # 既存のデータをクリア (1行目から最大行まで削除)
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    all_rows_data = []
    
    # pdfplumberでPDFを開き、全ページの全テーブルを抽出
    with pdfplumber.open(pdf_bytes_io) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    # tableはネストされたリスト (list of lists)
                    all_rows_data.extend(table)
    
    # 抽出したデータをシートに書き込む
    if not all_rows_data:
        ws.cell(row=1, column=1, value="PDFからテーブルデータを抽出できませんでした。")
    else:
        for r_idx, row_data in enumerate(all_rows_data, start=1):
            if row_data: # 行データが空でないことを確認
                for c_idx, cell_data in enumerate(row_data, start=1):
                    # セルデータがNoneの場合は空文字に変換
                    cell_value = cell_data if cell_data is not None else ""
                    ws.cell(row=r_idx, column=c_idx, value=cell_value)

    # 変更をバイトデータとして保存
    output_excel = io.BytesIO()
    wb.save(output_excel)
    return output_excel.getvalue()
# --- シール/その他PDF処理関数ここまで ---


if 'master_df' not in st.session_state:
    st.session_state.master_df = load_master_data("商品マスタ一覧", ['商品予定名', 'パン箱入数', '商品名', '売価単価', '弁当区分'])
if 'customer_master_df' not in st.session_state:
    st.session_state.customer_master_df = load_master_data("得意先マスタ一覧", ['得意先ＣＤ', '得意先名'])

st.markdown("""
    <style>
        [data-testid="stSidebarNav"] ul { display: none; }
        .custom-title {
            font-size: 2.1rem; font-weight: 600; color: #3A322E;
            padding-bottom: 10px; border-bottom: 3px solid #FF9933; margin-bottom: 25px;
        }
        .stApp { background: #fff5e6; }
    </style>
""", unsafe_allow_html=True)

st.sidebar.title("メニュー")
st.sidebar.page_link("streamlit_app.py", label="PDF Excel 変換", icon="📄")
st.sidebar.page_link("pages/マスタ設定.py", label="マスタ設定", icon="⚙️")
st.markdown('<p class="custom-title">数出表 PDF変換ツール</p>', unsafe_allow_html=True)
show_debug = st.sidebar.checkbox("デバッグ情報を表示", value=False)
uploaded_pdf = st.file_uploader("処理するPDFファイルをアップロードしてください", type="pdf", label_visibility="collapsed")

if uploaded_pdf is not None:
    
    pdf_bytes_io = io.BytesIO(uploaded_pdf.getvalue())
    original_pdf_name = os.path.splitext(uploaded_pdf.name)[0]
    
    df_paste_sheet = None
    is_suudashihyou = False

    # --- PDF種別判定 ---
    try:
        with st.spinner("PDFを分析中... (数出表か確認)"):
            # まず「数出表」としての処理を試みる
            df_paste_sheet = pdf_to_excel_data_for_paste_sheet(io.BytesIO(pdf_bytes_io.getvalue()))
        
        # 抽出が成功し、データが空でなければ「数出表」とみなす
        if df_paste_sheet is not None and not df_paste_sheet.empty:
            is_suudashihyou = True
        else:
            if show_debug: st.info("数出表の形式と一致しませんでした。シール/その他PDFとして処理します。")
    
    except Exception as e:
        # 抽出処理自体がエラーになった場合も「数出表」ではないとみなす
        is_suudashihyou = False
        if show_debug:
            st.info("数出表の解析に失敗しました。シール/その他PDFとして処理します。")
            st.warning(f"解析エラー: {str(e)}")

    # --- 処理分岐 ---
    if is_suudashihyou:
        # --- (A) 数出表PDFの処理 (既存のロジック) ---
        st.info("📊 数出表PDFとして処理します...")
        template_path = "template.xlsm"
        nouhinsyo_path = "nouhinsyo.xlsx"
        if not os.path.exists(template_path) or not os.path.exists(nouhinsyo_path):
            st.error(f"必要なテンプレートファイルが見つかりません：'{template_path}' または '{nouhinsyo_path}'")
            st.stop()
        
        template_wb = load_workbook(template_path, keep_vba=True)
        nouhinsyo_wb = load_workbook(nouhinsyo_path)
        
        # マスタデータをExcelに書き込む (既存の処理)
        try:
            df_product_master = load_master_csv("商品マスタ")
            if not df_product_master.empty and "商品マスタ" in template_wb.sheetnames:
                ws_product = template_wb["商品マスタ"]
                for row in ws_product.iter_rows():
                    for cell in row: cell.value = None
                paste_dataframe_to_sheet(ws_product, df_product_master)
                if show_debug: st.write("✅ 商品マスタを template.xlsm に貼り付けました")
        except Exception as e:
            if show_debug: st.warning(f"商品マスタの貼り付けエラー: {str(e)}")
        
        try:
            df_customer_master = load_master_csv("得意先マスタ")
            if not df_customer_master.empty and "得意先マスタ" in template_wb.sheetnames:
                ws_customer = template_wb["得意先マスタ"]
                for row in ws_customer.iter_rows():
                    for cell in row: cell.value = None
                paste_dataframe_to_sheet(ws_customer, df_customer_master)
                if show_debug: st.write("✅ 得意先マスタを template.xlsm に貼り付けました")
        except Exception as e:
            if show_debug: st.warning(f"得意先マスタの貼り付けエラー: {str(e)}")
        
        # 弁当・クライアントデータの抽出
        df_bento_sheet, df_client_sheet = None, None
        with st.spinner("PDFから詳細データを抽出中..."):
            try:
                tables = extract_table_from_pdf_for_bento(io.BytesIO(pdf_bytes_io.getvalue()))
                if tables:
                    main_table = max(tables, key=len)
                    anchor_col = find_correct_anchor_for_bento(main_table)
                    if anchor_col != -1:
                        bento_list = extract_bento_range_for_bento(main_table, anchor_col)
                        if bento_list:
                            matched_data = match_bento_data(bento_list, st.session_state.master_df)
                            df_bento_sheet = pd.DataFrame(matched_data, columns=['商品予定名', 'パン箱入数', '売価単価', '弁当区分'])
                            if show_debug:
                                st.write("--- 抽出・マッチング後の最終データ ---")
                                st.dataframe(df_bento_sheet)
            except Exception as e:
                st.error(f"注文弁当データ処理中にエラーが発生しました: {str(e)}")
                if show_debug: st.exception(e)

            try:
                client_data = extract_detailed_client_info_from_pdf(io.BytesIO(pdf_bytes_io.getvalue()))
                if client_data:
                    df_client_sheet = export_detailed_client_data_to_dataframe(client_data)
            except Exception as e:
                st.error(f"クライアント情報抽出中にエラーが発生しました: {str(e)}")
        
        # Excelファイルの生成 (既存の処理)
        if df_paste_sheet is not None:
            try:
                with st.spinner("Excelファイルを作成中..."):
                    # template.xlsm (マクロ有効) の作成
                    ws_paste = template_wb["貼り付け用"]
                    for r_idx, row in df_paste_sheet.iterrows():
                        for c_idx, value in enumerate(row):
                            ws_paste.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    if df_bento_sheet is not None:
                        safe_write_df(template_wb["注文弁当の抽出"], df_bento_sheet, start_row=1)
                    if df_client_sheet is not None:
                        safe_write_df(template_wb["クライアント抽出"], df_client_sheet, start_row=1)
                    
                    output_macro = io.BytesIO()
                    template_wb.save(output_macro)
                    macro_excel_bytes = output_macro.getvalue()

                    # nouhinsyo.xlsx (データのみ) の作成
                    df_bento_for_nouhin = None
                    if df_bento_sheet is not None:
                        master_df = st.session_state.master_df.copy()
                        master_df.columns = master_df.columns.str.strip()
                        if not master_df.empty and '商品名' in master_df.columns:
                            master_map = master_df.drop_duplicates(subset=['商品予定名']).set_index('商品予定名')['商品名'].to_dict()
                            df_bento_for_nouhin = df_bento_sheet.copy()
                            df_bento_for_nouhin['商品名'] = df_bento_for_nouhin['商品予定名'].map(master_map)
                            df_bento_for_nouhin = df_bento_for_nouhin[['商品予定名', 'パン箱入数', '商品名']]
                    
                    ws_paste_n = nouhinsyo_wb["貼り付け用"]
                    for r_idx, row in df_paste_sheet.iterrows():
                        for c_idx, value in enumerate(row):
                            ws_paste_n.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    if df_bento_for_nouhin is not None:
                        safe_write_df(nouhinsyo_wb["注文弁当の抽出"], df_bento_for_nouhin, start_row=1)
                    if df_client_sheet is not None:
                        safe_write_df(nouhinsyo_wb["クライアント抽出"], df_client_sheet, start_row=1)
                    if not st.session_state.customer_master_df.empty:
                        safe_write_df(nouhinsyo_wb["得意先マスタ"], st.session_state.customer_master_df, start_row=1)
                    
                    output_data_only = io.BytesIO()
                    nouhinsyo_wb.save(output_data_only)
                    data_only_excel_bytes = output_data_only.getvalue()

                st.success("✅ ファイルの準備が完了しました！")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="▼　数出表ダウンロード", data=macro_excel_bytes,
                        file_name=f"{original_pdf_name}_数出表.xlsm",
                        mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                    )
                with col2:
                    st.download_button(
                        label="▼　納品書ダウンロード", data=data_only_excel_bytes,
                        file_name=f"{original_pdf_name}_納品書.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Excelファイル生成中にエラーが発生しました: {str(e)}")

    else:
        # --- (B) シール/その他PDFの処理 (新規のロジック) ---
        st.info("📜 シール/その他PDFとして処理します...")
        seal_template_path = "seal.xlsx" # フォルダにある既存のファイル
        
        if not os.path.exists(seal_template_path):
            st.error(f"必要なテンプレートファイル '{seal_template_path}' が見つかりません。")
            st.stop()
        
        try:
            with st.spinner(f"'{seal_template_path}' にPDFデータを書き込み中..."):
                # 新しい関数を呼び出し、変更されたExcelのバイトデータを取得
                modified_seal_bytes = process_other_pdf_to_seal_template(pdf_bytes_io, seal_template_path)
            
            st.success(f"✅ 処理が完了しました！")
            st.download_button(
                label=f"▼ {seal_template_path} ダウンロード",
                data=modified_seal_bytes,
                file_name="seal.xlsx", # ファイル名は 'seal.xlsx' のまま
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"シール/その他PDF処理中にエラーが発生しました: {str(e)}")
            if show_debug:
                st.exception(e)
